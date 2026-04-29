"""Geração de planilha XLSX do projeto.

Saídas em 4 abas:
  - Resumo: metadados do projeto + métricas
  - Top N (incluídos no recorte final, ordenados por Quality Score)
  - Incluídos fora do Top
  - Excluídos
"""

from __future__ import annotations

import io
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(start_color="FFEC4899", end_color="FFEC4899", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
_SECTION_FILL = PatternFill(start_color="FFF1F5F9", end_color="FFF1F5F9", fill_type="solid")
_SECTION_FONT = Font(bold=True, color="FF0F172A", size=12)


_COLUMNS = [
    ("rank", "Rank", 6),
    ("quality_score", "Quality Score", 14),
    ("decision", "Decisão", 10),
    ("source", "Fonte", 10),
    ("year", "Ano", 7),
    ("title", "Título", 60),
    ("authors", "Autores", 40),
    ("journal", "Revista/Periódico", 30),
    ("doi", "DOI", 22),
    ("external_url", "Link", 40),
    ("summary_pt", "Resumo (IA)", 60),
    ("reason", "Motivo da decisão", 50),
]


def _write_sheet(ws, title: str, rows: Iterable[dict[str, Any]],
                 *, decision_label: str | None = None) -> None:
    ws.title = title[:31]  # Excel limit
    # Cabeçalho
    for col_idx, (_key, header, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Linhas
    for i, r in enumerate(rows, start=1):
        row_idx = i + 1
        score = r.get("quality_score")
        decision_val = decision_label or r.get("decision") or ""
        values = [
            i,
            score if score is not None else "",
            decision_val,
            r.get("source") or "",
            r.get("year") or "",
            r.get("title") or "",
            r.get("authors") or "",
            r.get("journal") or "",
            r.get("doi") or "",
            r.get("external_url") or "",
            r.get("summary_pt") or "",
            r.get("reason") or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx == 2 and isinstance(val, int):
                cell.alignment = Alignment(horizontal="center", vertical="top")
        ws.row_dimensions[row_idx].height = 60


def _write_resumo(ws, project, metrics: dict[str, Any]) -> None:
    ws.title = "Resumo"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 80

    rows: list[tuple[str, Any]] = [
        ("Tema", project.topic or ""),
        ("Objetivo", project.objective or ""),
        ("Tipo de revisão",
         "Revisão narrativa" if getattr(project, "review_type", "") == "narrative_review"
         else "Revisão sistemática"),
        ("Janela temporal", f"Últimos {project.years_window or 10} anos"),
        ("Meta de Top N", project.target_articles),
        ("Status", project.status),
        ("Criado em", project.created_at or ""),
        ("", ""),
        ("Total artigos", metrics.get("total_articles", 0)),
        ("Incluídos", metrics.get("included", 0)),
        ("Excluídos", metrics.get("excluded", 0)),
        ("Top final (incluídos)", metrics.get("included_top", 0)),
        ("Quality Score médio (Top)", metrics.get("top40_avg") or "—"),
    ]

    title_cell = ws.cell(row=1, column=1, value="Relatório do projeto")
    title_cell.font = Font(bold=True, size=14, color="FFEC4899")
    ws.cell(row=1, column=2, value=f"#{project.id}")
    ws.row_dimensions[1].height = 24

    for i, (label, value) in enumerate(rows, start=3):
        a = ws.cell(row=i, column=1, value=label)
        b = ws.cell(row=i, column=2, value=value)
        if label:
            a.font = Font(bold=True, color="FF334155")
            a.fill = _SECTION_FILL
        b.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A3"


def generate_project_xlsx(*, project, metrics: dict[str, Any],
                           top_n: list[dict], below_cutoff: list[dict] | None,
                           excluded: list[dict]) -> bytes:
    """Gera arquivo XLSX do projeto. Retorna bytes prontos pra Response."""
    wb = Workbook()
    # Aba 1: Resumo (substitui a default)
    _write_resumo(wb.active, project, metrics)

    # Aba 2: Top N
    ws_top = wb.create_sheet()
    _write_sheet(ws_top, f"Top {project.target_articles}",
                 top_n, decision_label="Incluído (Top)")

    # Aba 3: incluídos fora do top (opcional)
    if below_cutoff:
        ws_below = wb.create_sheet()
        _write_sheet(ws_below, "Incluidos fora do Top",
                     below_cutoff, decision_label="Incluído (fora do Top)")

    # Aba 4: excluídos
    ws_ex = wb.create_sheet()
    _write_sheet(ws_ex, "Excluidos", excluded, decision_label="Excluído")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
